# -*- coding: utf-8 -*-
"""
Created on Wed Nov 25 16:31:14 2020

@author: furqa
"""

from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import filedialog
import threading

root=Tk()
root.title("Forever Traders")
root.geometry("300x300")

wbA =Workbook()

wbforsave=Workbook()
wbforsave

#lists
AFileList=[]


def A():
    global fileA
    fileA=filedialog.askopenfilename()
    wbA=load_workbook(fileA)
    ws=wbA.active
    cA=ws['A']
    global FilesDestination
    FilesDestination= filedialog.askdirectory()
    for cell in cA:  
        if(cell.value!=None):
            x=FilesDestination+'/'+cell.value+'.xlsx'
            wbforsave.save(x)
            AFileList.append(cell.value)
    

    
def B():
    fileB=filedialog.askopenfilename()
    wbB=load_workbook(fileB)
    wsB=wbB.active
    #second file max col and rows
    maxRows=wsB.max_row
    maxCol=wsB.max_column
    t1=threading.Thread(target=DispProcessing ,args=["Processing...",0,2])
    t1.start()
    i:int
    for k in AFileList:
        y=FilesDestination+'/'+k+'.xlsx'
        wbindex=load_workbook(y)
        wsindex=wbindex.active
        
        for i in range(1,maxRows+1):
            for j in range(1,maxCol+1):
                if i==1 and j==1:
                    z=wsB.cell(row=i,column=j)
                    wsindex.cell(row=i,column=j).value=z.value
                elif j==1:
                    z=wsB.cell(row=i,column=j)
                    wsindex.cell(row=i,column=j).value=k+str(z.value)
                    #print(wsindex.cell(row=i,column=j).value)
                    
                else:
                    z=wsB.cell(row=i,column=j)
                    wsindex.cell(row=i,column=j).value=z.value
                    #print(wsindex.cell(row=i,column=j).value)
        wbindex.save(y)
    DispProcessing("Processing Completed!", 0,3)
        
def DispProcessing(StrPros,x,RV):
  if x==0:
      FT=Label(root, 
      		 text=StrPros,
      		 foreground = "White",
      		 background = "#2B2D2F",
      		 font = "Helvetica 16").grid(row=RV,column=0,padx=40,pady=20)

        

#buttonsA
ButtonA=Button(root,text="Choose First File and Destination for Files",command=A,border=1)
ButtonA.grid(row=0,column=0,padx=40,pady=20)


#buttonsB
ButtonB=Button(root,text="Choose Second File",command=B)
ButtonB.grid(row=1,column=0,padx=50,pady=20)

ProcessLabel=Label(root,text="Processing...")



root.mainloop()